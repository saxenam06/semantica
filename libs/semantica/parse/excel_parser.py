"""
Excel document parser for Semantica framework.

This module handles Excel file parsing using openpyxl and pandas
for spreadsheet data extraction.
"""

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import pandas as pd
from openpyxl import load_workbook

from ..utils.exceptions import ProcessingError, ValidationError
from ..utils.logging import get_logger


@dataclass
class ExcelSheet:
    """Excel sheet representation."""
    
    name: str
    data: List[Dict[str, Any]]
    row_count: int
    column_count: int
    headers: List[str] = field(default_factory=list)


@dataclass
class ExcelData:
    """Excel workbook representation."""
    
    sheet_names: List[str]
    sheets: Dict[str, ExcelSheet]
    metadata: Dict[str, Any] = field(default_factory=dict)


class ExcelParser:
    """Excel document parser."""
    
    def __init__(self, **config):
        """
        Initialize Excel parser.
        
        Args:
            **config: Parser configuration
        """
        self.logger = get_logger("excel_parser")
        self.config = config
    
    def parse(
        self,
        file_path: Union[str, Path],
        sheet_name: Optional[str] = None,
        **options
    ) -> Union[ExcelData, ExcelSheet]:
        """
        Parse Excel file.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Specific sheet name (None = all sheets)
            **options: Parsing options:
                - header_row: Row number to use as header (default: 0)
                - skip_rows: Number of rows to skip
                - max_rows: Maximum number of rows to read per sheet
                - engine: Parser engine ('pandas' or 'openpyxl', default: 'pandas')
                
        Returns:
            ExcelData or ExcelSheet: Parsed Excel data
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise ValidationError(f"Excel file not found: {file_path}")
        
        if not file_path.suffix.lower() in ['.xlsx', '.xls', '.xlsm']:
            raise ValidationError(f"File is not an Excel file: {file_path}")
        
        try:
            engine = options.get("engine", "pandas")
            
            if engine == "pandas":
                return self._parse_with_pandas(file_path, sheet_name, **options)
            else:
                return self._parse_with_openpyxl(file_path, sheet_name, **options)
                
        except Exception as e:
            self.logger.error(f"Failed to parse Excel {file_path}: {e}")
            raise ProcessingError(f"Failed to parse Excel: {e}")
    
    def _parse_with_pandas(
        self,
        file_path: Path,
        sheet_name: Optional[str],
        **options
    ) -> Union[ExcelData, ExcelSheet]:
        """Parse Excel using pandas."""
        header_row = options.get("header_row", 0)
        skip_rows = options.get("skip_rows", 0)
        max_rows = options.get("max_rows")
        
        # Read all sheets or specific sheet
        if sheet_name:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=header_row,
                skiprows=skip_rows,
                nrows=max_rows
            )
            
            # Convert to list of dictionaries
            rows = df.fillna("").to_dict('records')
            
            return ExcelSheet(
                name=sheet_name,
                data=rows,
                row_count=len(rows),
                column_count=len(df.columns),
                headers=list(df.columns)
            )
        else:
            # Read all sheets
            excel_file = pd.ExcelFile(file_path)
            sheets = {}
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(
                    excel_file,
                    sheet_name=sheet_name,
                    header=header_row,
                    skiprows=skip_rows,
                    nrows=max_rows
                )
                
                rows = df.fillna("").to_dict('records')
                
                sheets[sheet_name] = ExcelSheet(
                    name=sheet_name,
                    data=rows,
                    row_count=len(rows),
                    column_count=len(df.columns),
                    headers=list(df.columns)
                )
            
            return ExcelData(
                sheet_names=excel_file.sheet_names,
                sheets=sheets,
                metadata={
                    "file_path": str(file_path),
                    "total_sheets": len(excel_file.sheet_names)
                }
            )
    
    def _parse_with_openpyxl(
        self,
        file_path: Path,
        sheet_name: Optional[str],
        **options
    ) -> Union[ExcelData, ExcelSheet]:
        """Parse Excel using openpyxl."""
        wb = load_workbook(file_path, data_only=True)
        
        header_row = options.get("header_row", 0)
        skip_rows = options.get("skip_rows", 0)
        max_rows = options.get("max_rows")
        
        if sheet_name:
            ws = wb[sheet_name]
            return self._parse_sheet(ws, header_row, skip_rows, max_rows)
        else:
            sheets = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheets[sheet_name] = self._parse_sheet(ws, header_row, skip_rows, max_rows)
            
            return ExcelData(
                sheet_names=wb.sheetnames,
                sheets=sheets,
                metadata={
                    "file_path": str(file_path),
                    "total_sheets": len(wb.sheetnames)
                }
            )
    
    def _parse_sheet(self, worksheet, header_row: int, skip_rows: int, max_rows: Optional[int]) -> ExcelSheet:
        """Parse individual worksheet."""
        rows_data = []
        
        # Get headers
        headers = []
        if worksheet.max_row > header_row:
            header_cells = list(worksheet.iter_rows(min_row=header_row+1, max_row=header_row+1, values_only=True))[0]
            headers = [str(cell) if cell is not None else f"Column_{i+1}" for i, cell in enumerate(header_cells)]
        
        # Get data rows
        start_row = header_row + skip_rows + 2 if headers else skip_rows + 1
        end_row = start_row + max_rows if max_rows else worksheet.max_row + 1
        
        for row in worksheet.iter_rows(min_row=start_row, max_row=end_row-1, values_only=True):
            row_dict = {}
            for idx, cell_value in enumerate(row):
                header = headers[idx] if idx < len(headers) else f"Column_{idx+1}"
                row_dict[header] = str(cell_value) if cell_value is not None else ""
            rows_data.append(row_dict)
        
        return ExcelSheet(
            name=worksheet.title,
            data=rows_data,
            row_count=len(rows_data),
            column_count=len(headers) if headers else 0,
            headers=headers
        )
    
    def extract_sheet(self, file_path: Union[str, Path], sheet_name: str, **options) -> ExcelSheet:
        """
        Extract specific sheet from Excel.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name to extract
            **options: Parsing options
            
        Returns:
            ExcelSheet: Extracted sheet data
        """
        result = self.parse(file_path, sheet_name=sheet_name, **options)
        if isinstance(result, ExcelSheet):
            return result
        else:
            if sheet_name in result.sheets:
                return result.sheets[sheet_name]
            else:
                raise ValidationError(f"Sheet '{sheet_name}' not found in Excel file")
